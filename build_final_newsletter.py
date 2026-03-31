from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path
from urllib.parse import urlsplit

from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook

from prompts import newsletter_batch_prompt

ROOT = Path(__file__).resolve().parent
load_dotenv(ROOT / ".env")

DEFAULT_EXCEL_PATH = Path("recent_sitemap_outputs") / "recent_urls.xlsx"
DEFAULT_OUTPUT_DIR = Path("recent_sitemap_outputs") / "final_newsletters"
FINAL_NEWSLETTER_FILENAME = "automobile_tech_newsletter.md"
DEFAULT_MODEL = os.getenv("NEWSLETTER_LLM_MODEL") or os.getenv("NEWS_AGENT_LLM_MODEL", "gpt-5.4-nano")
NEWSLETTER_FOCUS = (
    "Keep automobile-technology news that has proper evidence, real technical depth, and meaningful relevance to "
    "breakthrough technology or significant automobile-technology issues. Reject only weak, repetitive, generic, "
    "or poorly supported items."
)
BATCH_SIZE = 5
URL_COLUMN_CANDIDATES = ("link", "url", "page_url", "article_url")
SELECTED_COLUMN_CANDIDATES = ("semantic_match",)
PROCESSED_PATH_COLUMN_CANDIDATES = ("processed_markdown_path",)
RELEVANCE_SCORE_COLUMN_CANDIDATES = ("relevance_score",)
MAX_ARTICLE_CHARS = 5000
BATCH_SELECTION_SCHEMA = {
    "name": "newsletter_batch_selection",
    "description": "Choose the strongest newsletter-worthy items from one batch of processed markdown files.",
    "parameters": {
        "type": "object",
        "properties": {
            "selected_items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "article_id": {
                            "type": "string",
                            "description": "One of the article IDs supplied in the batch.",
                        },
                        "newsletter_title": {
                            "type": "string",
                            "description": "Tighter headline for the final newsletter.",
                        },
                        "why_keep": {
                            "type": "string",
                            "description": "Why this item deserves inclusion in the final newsletter.",
                        },
                        "evidence_points": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Concrete technical or evidence-based points supporting inclusion.",
                        },
                    },
                    "required": ["article_id", "newsletter_title", "why_keep", "evidence_points"],
                    "additionalProperties": False,
                },
            }
        },
        "required": ["selected_items"],
        "additionalProperties": False,
    },
}

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a final technical automobile newsletter from selected processed markdown files sheet by sheet in batches of 5."
    )
    parser.add_argument(
        "--excel-path",
        default=str(DEFAULT_EXCEL_PATH),
        help="Path to the Excel workbook. Defaults to recent_sitemap_outputs/recent_urls.xlsx.",
    )
    parser.add_argument("--sheet-name", help="Process only one sheet. If omitted, all sheets are processed.")
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Directory for the generated final newsletter markdown file.",
    )
    return parser.parse_args()


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def parse_relevance_score(value: object) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def read_header_map(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, worksheet.max_column + 1):
        header = normalize_header(worksheet.cell(row=1, column=col).value)
        if header:
            headers[header] = col
    return headers


def load_workbook_and_sheets(excel_path: Path, sheet_name: str | None):
    workbook = load_workbook(excel_path, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' was not found. Available sheets: {workbook.sheetnames}")
        return workbook, [workbook[sheet_name]]
    return workbook, list(workbook.worksheets)


def detect_column(
    headers: dict[str, int],
    candidates: tuple[str, ...],
    contains_groups: tuple[tuple[str, ...], ...],
    label: str,
) -> str:
    for candidate in candidates:
        if candidate in headers:
            return candidate

    for header in headers:
        if any(all(token in header for token in token_group) for token_group in contains_groups):
            return header

    raise ValueError(f"Could not auto-detect a {label}. Available headers: {sorted(headers)}")


def resolve_output_dir(output_dir_arg: str) -> Path:
    output_dir = Path(output_dir_arg)
    if not output_dir.is_absolute():
        output_dir = (ROOT / output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def resolve_existing_file(excel_path: Path, path_value: object) -> Path | None:
    relative_path = normalize_text(path_value)
    if not relative_path:
        return None
    file_path = (excel_path.parent / relative_path).resolve()
    return file_path if file_path.exists() else None


def compact_text(markdown_text: str, max_chars: int) -> str:
    cleaned = re.sub(r"\n{3,}", "\n\n", markdown_text.strip())
    return cleaned if len(cleaned) <= max_chars else cleaned[:max_chars].rstrip()


def make_safe_slug(value: str, fallback: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
    return (slug or fallback)[:120]


def extract_title(markdown_text: str, fallback: str) -> str:
    for line in markdown_text.splitlines():
        stripped = line.strip()
        if stripped.startswith("# "):
            return stripped[2:].strip()
    for line in markdown_text.splitlines():
        stripped = line.strip()
        if stripped:
            return stripped[:160]
    return fallback


def chunk_items(items: list[dict[str, object]], size: int) -> list[list[dict[str, object]]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def init_model(
    *,
    api_key: str | None = None,
    model_name: str | None = None,
    base_url: str | None = None,
) -> ChatOpenAI:
    resolved_api_key = normalize_text(api_key or os.getenv("NEWS_AGENT_OPENAI_API_KEY") or os.getenv("OPENAI_API_KEY"))
    if not resolved_api_key:
        raise EnvironmentError("Set OPENAI_API_KEY or NEWS_AGENT_OPENAI_API_KEY before running build_final_newsletter.py.")

    resolved_model = normalize_text(model_name or DEFAULT_MODEL) or DEFAULT_MODEL
    resolved_base_url = normalize_text(base_url or os.getenv("NEWS_AGENT_OPENAI_BASE_URL") or os.getenv("OPENAI_BASE_URL"))
    model_kwargs = {
        "model": resolved_model,
        "api_key": resolved_api_key,
        "max_retries": 2,
        "timeout": 120,
    }
    if resolved_base_url:
        model_kwargs["base_url"] = resolved_base_url
    return ChatOpenAI(**model_kwargs)


def build_batch_chain(model: ChatOpenAI):
    return newsletter_batch_prompt | model.with_structured_output(
        BATCH_SELECTION_SCHEMA,
        method="function_calling",
    )


def collect_selected_articles(excel_path: Path, worksheet) -> list[dict[str, object]]:
    headers = read_header_map(worksheet)
    url_key = detect_column(headers, URL_COLUMN_CANDIDATES, contains_groups=(("url",), ("link",)), label="URL column")
    selected_key = detect_column(
        headers,
        SELECTED_COLUMN_CANDIDATES,
        contains_groups=(("semantic", "match"),),
        label="selection status column",
    )
    processed_key = detect_column(
        headers,
        PROCESSED_PATH_COLUMN_CANDIDATES,
        contains_groups=(("processed", "markdown", "path"),),
        label="processed markdown path column",
    )
    relevance_key = detect_column(
        headers,
        RELEVANCE_SCORE_COLUMN_CANDIDATES,
        contains_groups=(("relevance", "score"),),
        label="relevance score column",
    )

    articles: list[dict[str, object]] = []
    for row_number in range(2, worksheet.max_row + 1):
        selected_status = normalize_text(worksheet.cell(row=row_number, column=headers[selected_key]).value).lower()
        if selected_status != "selected":
            continue

        source_url = normalize_text(worksheet.cell(row=row_number, column=headers[url_key]).value)
        processed_path = resolve_existing_file(excel_path, worksheet.cell(row=row_number, column=headers[processed_key]).value)
        if processed_path is None:
            continue

        markdown_text = processed_path.read_text(encoding="utf-8")
        title_fallback = make_safe_slug(urlsplit(source_url).path, "article").replace("_", " ")
        articles.append(
            {
                "article_id": f"{worksheet.title}-row-{row_number:06d}",
                "sheet_name": worksheet.title,
                "row_number": row_number,
                "source_url": source_url,
                "processed_markdown_path": processed_path,
                "relevance_score": parse_relevance_score(
                    worksheet.cell(row=row_number, column=headers[relevance_key]).value
                ),
                "title": extract_title(markdown_text, title_fallback),
                "markdown_text": compact_text(markdown_text, MAX_ARTICLE_CHARS),
            }
        )

    articles.sort(key=lambda item: (-int(item["relevance_score"]), int(item["row_number"])))
    return articles


def format_batch_articles(batch_articles: list[dict[str, object]]) -> str:
    blocks = []
    for article in batch_articles:
        blocks.append(
            "\n".join(
                [
                    f'Article ID: {article["article_id"]}',
                    f'Title: {article["title"]}',
                    f'Source URL: {article["source_url"]}',
                    f'Relevance Score: {article["relevance_score"]}',
                    "Processed markdown:",
                    str(article["markdown_text"]),
                ]
            )
        )
    return "\n\n---\n\n".join(blocks)


def shortlist_batch(chain, sheet_name: str, batch_articles: list[dict[str, object]]) -> list[dict[str, object]]:
    response = chain.invoke(
        {
            "newsletter_focus": NEWSLETTER_FOCUS,
            "sheet_name": sheet_name,
            "batch_articles": format_batch_articles(batch_articles),
        }
    )
    if not isinstance(response, dict):
        raise RuntimeError("Batch selection returned an unexpected response type.")

    batch_map = {str(article["article_id"]): article for article in batch_articles}
    shortlisted: list[dict[str, object]] = []
    for item in response.get("selected_items", []):
        article_id = normalize_text(item.get("article_id"))
        if article_id not in batch_map:
            continue
        article = dict(batch_map[article_id])
        article["newsletter_title"] = normalize_text(item.get("newsletter_title")) or str(article["title"])
        article["why_keep"] = normalize_text(item.get("why_keep"))
        article["evidence_points"] = [normalize_text(point) for point in item.get("evidence_points", []) if normalize_text(point)]
        shortlisted.append(article)
    return shortlisted


def format_sheet_section(shortlisted_items: list[dict[str, object]], sheet_name: str) -> str:
    lines = [f"## {sheet_name}", ""]
    if not shortlisted_items:
        lines.append("No items were strong enough for the final newsletter.")
        return "\n".join(lines)

    for item in shortlisted_items:
        lines.append(f'### {item["newsletter_title"]}')
        evidence_points = item.get("evidence_points", [])
        if evidence_points:
            for point in evidence_points[:5]:
                lines.append(f"- {point}")
        else:
            fallback_points = [line.strip("- ").strip() for line in str(item["markdown_text"]).splitlines() if line.strip().startswith("- ")]
            if fallback_points:
                for point in fallback_points[:5]:
                    lines.append(f"- {point}")
            else:
                lines.append(f"- {item['why_keep']}")
        lines.append(f"Source: {item['source_url']}")
        lines.append("")
    return "\n".join(lines).rstrip()


def build_final_document(sheet_sections: list[str]) -> str:
    if not sheet_sections:
        raise RuntimeError("No sheet sections were generated for the final newsletter.")

    title = "# Automotive Tech Newsletter - Breakthroughs & Technical Milestones"
    editor_note = (
        "*Editor's note: This edition keeps only items with strong evidence, deep technical grounding, "
        "and clear breakthrough or significant automobile-technology relevance.*"
    )
    body = "\n\n---\n\n".join(section.strip() for section in sheet_sections if section.strip())
    return f"{title}\n{editor_note}\n\n---\n\n{body}\n"


def save_final_newsletter(output_dir: Path, final_markdown: str) -> Path:
    for old_file in output_dir.glob("automobile_tech_newsletter_*.md"):
        old_file.unlink(missing_ok=True)

    output_path = output_dir / FINAL_NEWSLETTER_FILENAME
    output_path.write_text(final_markdown, encoding="utf-8")
    return output_path


def run_newsletter_stage(
    *,
    excel_path: Path | str = DEFAULT_EXCEL_PATH,
    sheet_name: str | None = None,
    output_dir: Path | str = DEFAULT_OUTPUT_DIR,
    api_key: str | None = None,
    model_name: str | None = None,
    base_url: str | None = None,
    force_rebuild: bool = True,
) -> bool:
    resolved_excel_path = Path(excel_path).resolve()
    if not resolved_excel_path.exists():
        raise FileNotFoundError(f"Excel workbook not found: {resolved_excel_path}")

    resolved_output_dir = resolve_output_dir(str(output_dir))
    output_path = resolved_output_dir / FINAL_NEWSLETTER_FILENAME
    if output_path.exists() and not force_rebuild:
        print(f"[SKIP] Final newsletter already exists: {output_path}")
        return False

    workbook, worksheets = load_workbook_and_sheets(resolved_excel_path, sheet_name)
    try:
        model = init_model(api_key=api_key, model_name=model_name, base_url=base_url)
        batch_chain = build_batch_chain(model)
        shortlisted_items: list[dict[str, object]] = []
        for worksheet in worksheets:
            articles = collect_selected_articles(resolved_excel_path, worksheet)
            if not articles:
                continue

            print(f"\n[SHEET] {worksheet.title} | selected_articles={len(articles)}")
            for batch_number, batch_articles in enumerate(chunk_items(articles, BATCH_SIZE), start=1):
                print(f"[BATCH] {worksheet.title} | batch={batch_number} | size={len(batch_articles)}")
                shortlisted_items.extend(shortlist_batch(batch_chain, worksheet.title, batch_articles))

        if not shortlisted_items:
            raise RuntimeError("No newsletter-worthy items were shortlisted from the selected processed markdown files.")

        sheet_sections: list[str] = []
        for worksheet in worksheets:
            sheet_items = [item for item in shortlisted_items if item["sheet_name"] == worksheet.title]
            if not sheet_items:
                continue
            sheet_sections.append(format_sheet_section(sheet_items, worksheet.title))

        final_markdown = build_final_document(sheet_sections)
        save_final_newsletter(resolved_output_dir, final_markdown)
        print(f"\n[DONE] Final newsletter saved to: {output_path}")
        print(f"[DONE] Shortlisted items used: {len(shortlisted_items)}")
        return True
    finally:
        workbook.close()


def main() -> None:
    args = parse_args()
    run_newsletter_stage(
        excel_path=args.excel_path,
        sheet_name=args.sheet_name,
        output_dir=args.output_dir,
        force_rebuild=True,
    )


if __name__ == "__main__":
    main()
