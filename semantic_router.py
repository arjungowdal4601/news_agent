from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path
from urllib.parse import urlsplit

from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook

from prompts import semantic_router_prompt

ROOT = Path(__file__).resolve().parent
load_dotenv(ROOT / ".env")

DEFAULT_USER_NEED = os.getenv("USER_NEED", "REPLACE THIS WITH WHAT YOU NEED")
DEFAULT_EXCEL_PATH = Path("recent_sitemap_outputs") / "recent_urls.xlsx"
DEFAULT_OUTPUT_DIR = "processed_selected_markdown"
DEFAULT_MODEL = os.getenv("NEWS_AGENT_LLM_MODEL", "gpt-5.4-nano")
URL_COLUMN_CANDIDATES = ("link", "url", "page_url", "article_url")
MARKDOWN_PATH_COLUMN_CANDIDATES = ("markdown_path", "scraped_markdown_path", "md_path")
OUTPUT_COLUMNS = [
    "relevance_score",
    "semantic_match",
    "processed_markdown_saved",
    "processed_markdown_path",
]
DONE_SELECTION_STATUSES = {"selected", "not_selected", "yes", "no"}
DONE_SAVE_STATUSES = {"saved", "not_saved", "yes", "no", "missing_source_markdown"}
MAX_MARKDOWN_CHARS = 24000
MAX_IMAGE_CONTEXT_CHARS = 180
IMAGE_MARKDOWN_RE = re.compile(r"!\[(?P<alt>[^\]]*)\]\((?P<url>https?://[^)\s]+)")
DECORATIVE_IMAGE_HINTS = (
    "logo",
    "linkedin",
    "join our linkedin group",
    "rss",
    "newsletter",
    "subscribe",
    "icon",
    "avatar",
    "login",
)
ROUTER_OUTPUT_SCHEMA = {
    "name": "semantic_router_result",
    "description": "Decide whether an article matches the user requirement and rewrite only the relevant content.",
    "parameters": {
        "type": "object",
        "properties": {
            "relevance_score": {
                "type": "integer",
                "description": "Numeric score from 0 to 100 showing how relevant the article is to the user need.",
                "minimum": 0,
                "maximum": 100,
            },
            "matches": {
                "type": "boolean",
                "description": "True when the article is meaningfully relevant to the user requirement.",
            },
            "processed_markdown": {
                "type": "string",
                "description": "Clean rewritten markdown containing only the relevant parts. Empty when there is no match.",
            },
            "selected_image_ids": {
                "type": "array",
                "description": "IDs from the provided image catalog that should remain in the processed markdown because they are relevant.",
                "items": {"type": "integer"},
            },
        },
        "required": ["relevance_score", "matches", "processed_markdown", "selected_image_ids"],
        "additionalProperties": False,
    },
}

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read markdown paths from an Excel workbook, select relevant articles with a small OpenAI model through LangChain, and save processed markdown back to disk."
    )
    parser.add_argument(
        "--excel-path",
        default=str(DEFAULT_EXCEL_PATH),
        help="Path to the Excel workbook. Defaults to recent_sitemap_outputs/recent_urls.xlsx.",
    )
    parser.add_argument("--sheet-name", help="Sheet name to process. If omitted, all sheets are processed.")
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for processed markdown files. Relative paths are resolved from the workbook folder.",
    )
    parser.add_argument(
        "--force-reprocess",
        action="store_true",
        help="Re-run rows even if they already have semantic routing results.",
    )
    return parser.parse_args()


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def read_header_map(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, worksheet.max_column + 1):
        header = normalize_header(worksheet.cell(row=1, column=col).value)
        if header:
            headers[header] = col
    return headers


def load_workbook_and_sheets(excel_path: Path, sheet_name: str | None):
    workbook = load_workbook(excel_path)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' was not found. Available sheets: {workbook.sheetnames}")
        return workbook, [workbook[sheet_name]]
    return workbook, list(workbook.worksheets)


def detect_column(
    headers: dict[str, int],
    candidates: tuple[str, ...],
    contains_groups: tuple[tuple[str, ...], ...],
    label: str,
) -> str:
    for candidate in candidates:
        if candidate in headers:
            return candidate

    for header in headers:
        if any(all(token in header for token in token_group) for token_group in contains_groups):
            return header

    raise ValueError(f"Could not auto-detect a {label}. Available headers: {sorted(headers)}")


def ensure_output_columns_exist(worksheet, headers: dict[str, int]) -> tuple[dict[str, int], bool]:
    next_col = worksheet.max_column + 1
    changed = False
    for column_name in OUTPUT_COLUMNS:
        key = normalize_header(column_name)
        if key in headers:
            continue
        worksheet.cell(row=1, column=next_col, value=column_name)
        headers[key] = next_col
        next_col += 1
        changed = True
    return headers, changed


def resolve_output_dir(excel_path: Path, output_dir_arg: str) -> Path:
    output_dir = Path(output_dir_arg)
    if not output_dir.is_absolute():
        output_dir = excel_path.parent / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir.resolve()


def relative_to_workbook(excel_path: Path, file_path: Path) -> str:
    return os.path.relpath(file_path, start=excel_path.parent).replace("\\", "/")


def resolve_existing_file(excel_path: Path, path_value: object) -> Path | None:
    relative_path = normalize_text(path_value)
    if not relative_path:
        return None
    file_path = (excel_path.parent / relative_path).resolve()
    return file_path if file_path.exists() else None


def make_safe_slug(value: str, fallback: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
    return (slug or fallback)[:120]


def compact_markdown(markdown_text: str, max_chars: int) -> str:
    cleaned = re.sub(r"\n{3,}", "\n\n", markdown_text.strip())
    if len(cleaned) <= max_chars:
        return cleaned

    marker = "\n\n[... middle omitted for routing ...]\n\n"
    chunk_size = max((max_chars - len(marker)) // 2, 1)
    return cleaned[:chunk_size] + marker + cleaned[-chunk_size:]


def is_decorative_image(url: str, alt_text: str, context_text: str) -> bool:
    combined = f"{alt_text} {context_text} {url}".lower()
    return any(hint in combined for hint in DECORATIVE_IMAGE_HINTS)


def extract_image_candidates(markdown_text: str) -> list[dict[str, str]]:
    candidates: list[dict[str, str]] = []
    seen_urls: set[str] = set()

    for line in markdown_text.splitlines():
        stripped_line = line.strip()
        if not stripped_line:
            continue

        for match in IMAGE_MARKDOWN_RE.finditer(stripped_line):
            image_url = normalize_text(match.group("url"))
            if not image_url or image_url in seen_urls:
                continue

            alt_text = normalize_text(match.group("alt"))
            trailing_context = normalize_text(stripped_line[match.end() :])
            leading_context = normalize_text(stripped_line[: match.start()])
            context_text = trailing_context or leading_context or alt_text
            context_text = context_text[:MAX_IMAGE_CONTEXT_CHARS]

            if is_decorative_image(image_url, alt_text, context_text):
                continue

            seen_urls.add(image_url)
            candidates.append(
                {
                    "id": str(len(candidates) + 1),
                    "alt_text": alt_text,
                    "url": image_url,
                    "context": context_text,
                }
            )

    return candidates


def format_image_catalog(image_candidates: list[dict[str, str]]) -> str:
    if not image_candidates:
        return "No relevant source images were detected."

    lines = []
    for image in image_candidates:
        lines.append(
            f'{image["id"]}. alt="{image["alt_text"] or "Image"}" url="{image["url"]}" context="{image["context"]}"'
        )
    return "\n".join(lines)


def strip_fences(text: str) -> str:
    cleaned = normalize_text(text)
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```[a-zA-Z0-9_-]*\n?", "", cleaned)
        cleaned = re.sub(r"\n?```$", "", cleaned)
    return cleaned.strip()


def ensure_user_need_is_configured(user_need: str | None = None) -> str:
    need = normalize_text(user_need or DEFAULT_USER_NEED)
    if not need or "REPLACE THIS WITH WHAT YOU NEED" in need.upper():
        raise ValueError("Update USER_NEED in .env before running the script.")
    return need


def init_model(
    *,
    api_key: str | None = None,
    model_name: str | None = None,
    base_url: str | None = None,
) -> ChatOpenAI:
    resolved_api_key = normalize_text(api_key or os.getenv("NEWS_AGENT_OPENAI_API_KEY") or os.getenv("OPENAI_API_KEY"))
    if not resolved_api_key:
        raise EnvironmentError("Set OPENAI_API_KEY or NEWS_AGENT_OPENAI_API_KEY before running semantic_router.py.")

    resolved_model = normalize_text(model_name or os.getenv("NEWS_AGENT_LLM_MODEL", DEFAULT_MODEL)) or DEFAULT_MODEL
    resolved_base_url = normalize_text(base_url or os.getenv("NEWS_AGENT_OPENAI_BASE_URL") or os.getenv("OPENAI_BASE_URL"))
    model_kwargs = {
        "model": resolved_model,
        "api_key": resolved_api_key,
        "max_retries": 2,
        "timeout": 120,
    }
    if resolved_base_url:
        model_kwargs["base_url"] = resolved_base_url
    return ChatOpenAI(**model_kwargs)


def build_router_chain(model: ChatOpenAI):
    return semantic_router_prompt | model.with_structured_output(
        ROUTER_OUTPUT_SCHEMA,
        method="function_calling",
    )


def run_router(
    chain,
    need: str,
    source_url: str,
    markdown_text: str,
    image_catalog: str,
) -> dict[str, object]:
    response = chain.invoke(
        {
            "need": need,
            "source_url": source_url,
            "image_catalog": image_catalog,
            "markdown_content": markdown_text,
        }
    )
    if not isinstance(response, dict):
        raise RuntimeError("The model returned an unexpected response type.")
    return response


def update_worksheet_row(worksheet, row_number: int, headers: dict[str, int], **values) -> None:
    for key, value in values.items():
        worksheet.cell(row=row_number, column=headers[normalize_header(key)], value=value)


def normalize_relevance_score(value: object) -> int | None:
    try:
        score = int(value)
    except (TypeError, ValueError):
        return None
    return max(0, min(100, score))


def build_processed_markdown_path(
    output_dir: Path,
    table_name: str,
    row_number: int,
    source_url: str,
    markdown_path: Path,
) -> Path:
    table_slug = make_safe_slug(table_name, "sheet")
    domain_slug = make_safe_slug(urlsplit(source_url).netloc.lower(), "unknown_domain")
    source_stem = make_safe_slug(markdown_path.stem, f"row_{row_number:06d}")
    target_dir = output_dir / table_slug / domain_slug
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir / f"{source_stem}_processed.md"


def save_processed_markdown(
    excel_path: Path,
    output_dir: Path,
    table_name: str,
    row_number: int,
    source_url: str,
    markdown_path: Path,
    processed_markdown: str,
) -> str:
    target_path = build_processed_markdown_path(output_dir, table_name, row_number, source_url, markdown_path)
    target_path.write_text(processed_markdown, encoding="utf-8")
    return relative_to_workbook(excel_path, target_path)


def append_selected_images(
    processed_markdown: str,
    image_candidates: list[dict[str, str]],
    selected_image_ids: object,
) -> str:
    if not isinstance(selected_image_ids, list):
        return processed_markdown

    images_by_id = {int(image["id"]): image for image in image_candidates}
    selected_images: list[dict[str, str]] = []
    seen_ids: set[int] = set()

    for raw_id in selected_image_ids:
        try:
            image_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if image_id in seen_ids or image_id not in images_by_id:
            continue
        seen_ids.add(image_id)
        selected_images.append(images_by_id[image_id])

    if not selected_images:
        return processed_markdown

    image_blocks = []
    for image in selected_images:
        alt_text = image["alt_text"] or "Relevant image"
        image_blocks.append(f'![{alt_text}]({image["url"]})')
        if image["context"] and image["context"] != alt_text:
            image_blocks.append(f'*Source context: {image["context"]}*')

    return processed_markdown.rstrip() + "\n\n## Relevant Images\n\n" + "\n\n".join(image_blocks)


def should_skip_row(
    force_reprocess: bool,
    excel_path: Path,
    semantic_match: object,
    processed_save_status: object,
    processed_markdown_path: object,
) -> bool:
    if force_reprocess:
        return False

    match_value = normalize_text(semantic_match).lower()
    save_status = normalize_text(processed_save_status).lower()

    if match_value not in DONE_SELECTION_STATUSES:
        return False
    if save_status in {"saved", "yes"}:
        return resolve_existing_file(excel_path, processed_markdown_path) is not None
    return save_status in DONE_SAVE_STATUSES


def print_summary(summary: dict[str, int]) -> None:
    print("\nSummary")
    print(f"rows checked: {summary['checked']}")
    print(f"matches saved: {summary['matched_yes']}")
    print(f"non-matches: {summary['matched_no']}")
    print(f"saved markdown files: {summary['saved']}")
    print(f"missing markdown path: {summary['missing_markdown_path']}")
    print(f"missing markdown file: {summary['missing_markdown_file']}")
    print(f"failed rows: {summary['failed']}")
    print(f"skipped existing rows: {summary['skipped_existing']}")


def run_route_stage(
    *,
    excel_path: Path | str = DEFAULT_EXCEL_PATH,
    sheet_name: str | None = None,
    output_dir: str = DEFAULT_OUTPUT_DIR,
    force_reprocess: bool = False,
    user_need: str | None = None,
    api_key: str | None = None,
    model_name: str | None = None,
    base_url: str | None = None,
) -> bool:
    resolved_excel_path = Path(excel_path).resolve()
    if not resolved_excel_path.exists():
        raise FileNotFoundError(f"Excel workbook not found: {resolved_excel_path}")

    need = ensure_user_need_is_configured(user_need)
    resolved_output_dir = resolve_output_dir(resolved_excel_path, output_dir)
    workbook, worksheets = load_workbook_and_sheets(resolved_excel_path, sheet_name)
    chain = build_router_chain(init_model(api_key=api_key, model_name=model_name, base_url=base_url))
    summary = {
        "checked": 0,
        "matched_yes": 0,
        "matched_no": 0,
        "saved": 0,
        "missing_markdown_path": 0,
        "missing_markdown_file": 0,
        "failed": 0,
        "skipped_existing": 0,
    }
    schema_changed = False

    for worksheet in worksheets:
        headers, added = ensure_output_columns_exist(worksheet, read_header_map(worksheet))
        schema_changed = schema_changed or added
        url_key = detect_column(
            headers,
            URL_COLUMN_CANDIDATES,
            contains_groups=(("url",), ("link",)),
            label="URL column",
        )
        markdown_key = detect_column(
            headers,
            MARKDOWN_PATH_COLUMN_CANDIDATES,
            contains_groups=(("markdown", "path"), ("md", "path")),
            label="markdown path column",
        )
        workbook.save(resolved_excel_path)
        print(f"\n[SHEET] {worksheet.title} | url_column={url_key} | markdown_column={markdown_key}")

        for row_number in range(2, worksheet.max_row + 1):
            url = normalize_text(worksheet.cell(row=row_number, column=headers[url_key]).value)
            if not url:
                continue

            summary["checked"] += 1
            semantic_match = worksheet.cell(row=row_number, column=headers["semantic_match"]).value
            processed_save_status = worksheet.cell(row=row_number, column=headers["processed_markdown_saved"]).value
            processed_markdown_path = worksheet.cell(row=row_number, column=headers["processed_markdown_path"]).value

            if should_skip_row(
                force_reprocess,
                resolved_excel_path,
                semantic_match,
                processed_save_status,
                processed_markdown_path,
            ):
                summary["skipped_existing"] += 1
                print(f"[SKIP] Row {row_number}: {url}")
                continue

            markdown_path_value = worksheet.cell(row=row_number, column=headers[markdown_key]).value
            if not normalize_text(markdown_path_value):
                update_worksheet_row(
                    worksheet,
                    row_number,
                    headers,
                    relevance_score="",
                    semantic_match="source_missing",
                    processed_markdown_saved="missing_source_markdown",
                    processed_markdown_path="",
                )
                summary["missing_markdown_path"] += 1
                workbook.save(resolved_excel_path)
                print(f"[MISS] Row {row_number}: {url} -> markdown_path is empty")
                continue

            source_markdown_path = resolve_existing_file(resolved_excel_path, markdown_path_value)
            if source_markdown_path is None:
                update_worksheet_row(
                    worksheet,
                    row_number,
                    headers,
                    relevance_score="",
                    semantic_match="source_missing",
                    processed_markdown_saved="missing_source_markdown",
                    processed_markdown_path="",
                )
                summary["missing_markdown_file"] += 1
                workbook.save(resolved_excel_path)
                print(f"[MISS] Row {row_number}: {url} -> markdown file not found")
                continue

            try:
                print(f"[ROUTE] Row {row_number}: {url}")
                markdown_text = source_markdown_path.read_text(encoding="utf-8")
                image_candidates = extract_image_candidates(markdown_text)
                routed = run_router(
                    chain,
                    need,
                    url,
                    compact_markdown(markdown_text, MAX_MARKDOWN_CHARS),
                    format_image_catalog(image_candidates),
                )
                relevance_score = normalize_relevance_score(routed.get("relevance_score"))
                if relevance_score is None:
                    raise RuntimeError("The model did not return a valid relevance_score.")
                matched = bool(routed.get("matches"))
                processed_markdown = strip_fences(routed.get("processed_markdown"))
                processed_markdown = append_selected_images(
                    processed_markdown,
                    image_candidates,
                    routed.get("selected_image_ids"),
                )

                if not matched:
                    update_worksheet_row(
                        worksheet,
                        row_number,
                        headers,
                        relevance_score=relevance_score,
                        semantic_match="not_selected",
                        processed_markdown_saved="not_saved",
                        processed_markdown_path="",
                    )
                    summary["matched_no"] += 1
                else:
                    if not processed_markdown:
                        raise RuntimeError("The model marked the article as relevant but returned empty processed markdown.")

                    saved_path = save_processed_markdown(
                        resolved_excel_path,
                        resolved_output_dir,
                        worksheet.title,
                        row_number,
                        url,
                        source_markdown_path,
                        processed_markdown,
                    )
                    update_worksheet_row(
                        worksheet,
                        row_number,
                        headers,
                        relevance_score=relevance_score,
                        semantic_match="selected",
                        processed_markdown_saved="saved",
                        processed_markdown_path=saved_path,
                    )
                    summary["matched_yes"] += 1
                    summary["saved"] += 1
            except Exception as exc:
                update_worksheet_row(
                    worksheet,
                    row_number,
                    headers,
                    relevance_score="",
                    semantic_match="error",
                    processed_markdown_saved="error",
                    processed_markdown_path="",
                )
                summary["failed"] += 1
                print(f"[FAIL] Row {row_number}: {url} -> {exc}")
            finally:
                workbook.save(resolved_excel_path)

    print_summary(summary)
    workbook.close()
    return schema_changed or any(
        summary[key] > 0
        for key in ("matched_yes", "matched_no", "saved", "missing_markdown_path", "missing_markdown_file", "failed")
    )


def main() -> None:
    args = parse_args()
    run_route_stage(
        excel_path=args.excel_path,
        sheet_name=args.sheet_name,
        output_dir=args.output_dir,
        force_reprocess=args.force_reprocess,
    )


if __name__ == "__main__":
    main()
