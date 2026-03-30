from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
import warnings
from datetime import datetime
from pathlib import Path
from urllib.parse import urlsplit

from requests.exceptions import RequestsDependencyWarning

warnings.filterwarnings("ignore", category=RequestsDependencyWarning)

from crawl4ai import (
    AsyncWebCrawler,
    BrowserConfig,
    CacheMode,
    CrawlerRunConfig,
    DefaultMarkdownGenerator,
)
from openpyxl import load_workbook

OUTPUT_COLUMNS = [
    "markdown_path",
    "scrape_status",
    "scraped_at",
    "scrape_error",
    "markdown_type",
]
DEFAULT_EXCEL_PATH = Path("recent_sitemap_outputs") / "recent_urls.xlsx"
URL_COLUMN_CANDIDATES = ("link", "url", "page_url", "article_url")

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read URLs from an Excel sheet, crawl them with Crawl4AI, save markdown files, and write results back to the workbook."
    )
    parser.add_argument(
        "--excel-path",
        default=str(DEFAULT_EXCEL_PATH),
        help="Path to the Excel workbook. Defaults to recent_sitemap_outputs/recent_urls.xlsx.",
    )
    parser.add_argument("--sheet-name", help="Sheet name to process. If omitted, all sheets are processed.")
    parser.add_argument(
        "--url-column",
        help="Header name of the URL column, for example: link or url. If omitted, the script auto-detects it.",
    )
    parser.add_argument(
        "--output-dir",
        default="scraped_markdown",
        help="Base directory for markdown files. Relative paths are resolved from the workbook folder.",
    )
    parser.add_argument(
        "--force-rescrape",
        action="store_true",
        help="Rescrape rows even if markdown_path already exists and the file is present.",
    )
    parser.add_argument(
        "--headless",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Run the browser in headless mode. Use --no-headless for visible browsing.",
    )
    return parser.parse_args()


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def read_header_map(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, worksheet.max_column + 1):
        header = normalize_header(worksheet.cell(row=1, column=col).value)
        if header:
            headers[header] = col
    return headers


def load_workbook_and_sheets(excel_path: Path, sheet_name: str | None):
    workbook = load_workbook(excel_path)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' was not found. Available sheets: {workbook.sheetnames}")
        return workbook, [workbook[sheet_name]]
    return workbook, list(workbook.worksheets)


def detect_url_column(headers: dict[str, int], url_column: str | None) -> str:
    if url_column:
        url_key = normalize_header(url_column)
        if url_key not in headers:
            raise ValueError(f"URL column '{url_column}' was not found. Available headers: {sorted(headers)}")
        return url_key

    for candidate in URL_COLUMN_CANDIDATES:
        if candidate in headers:
            return candidate

    for header in headers:
        if "url" in header or "link" in header:
            return header

    raise ValueError(f"Could not auto-detect a URL column. Available headers: {sorted(headers)}")


def ensure_output_columns_exist(worksheet, headers: dict[str, int]) -> dict[str, int]:
    next_col = worksheet.max_column + 1
    for column_name in OUTPUT_COLUMNS:
        key = normalize_header(column_name)
        if key in headers:
            continue
        worksheet.cell(row=1, column=next_col, value=column_name)
        headers[key] = next_col
        next_col += 1
    return headers


def resolve_output_dir(excel_path: Path, output_dir_arg: str) -> Path:
    output_dir = Path(output_dir_arg)
    if not output_dir.is_absolute():
        output_dir = excel_path.parent / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir.resolve()


def make_safe_slug(value: str, fallback: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
    return (slug or fallback)[:120]


def slugify_filename_from_url(url: str, row_number: int) -> tuple[str, str]:
    parsed = urlsplit(url)
    domain = make_safe_slug(parsed.netloc.lower(), "unknown_domain")
    path_part = parsed.path.strip("/") or "home"
    if parsed.query:
        path_part = f"{path_part}_{parsed.query}"
    file_slug = make_safe_slug(path_part, "page")
    return domain, f"row_{row_number:06d}_{file_slug}.md"


def existing_markdown_file(excel_path: Path, relative_path: str) -> Path | None:
    relative_path = str(relative_path or "").strip()
    if not relative_path:
        return None
    file_path = (excel_path.parent / relative_path).resolve()
    return file_path if file_path.exists() else None


def get_raw_markdown(result) -> tuple[str, str]:
    markdown_value = getattr(result, "markdown", None)
    if markdown_value is None:
        raise RuntimeError("Crawl succeeded but no markdown was returned.")

    if hasattr(markdown_value, "raw_markdown"):
        return str(markdown_value.raw_markdown or ""), "raw_markdown"

    if isinstance(markdown_value, str):
        return markdown_value, "string_markdown"

    return str(markdown_value), type(markdown_value).__name__


def build_run_config() -> CrawlerRunConfig:
    return CrawlerRunConfig(
        cache_mode=CacheMode.BYPASS,
        markdown_generator=DefaultMarkdownGenerator(),
        page_timeout=60000,
        delay_before_return_html=1.0,
        wait_until="domcontentloaded",
        wait_for=None,
    )


async def crawl_one_url(crawler: AsyncWebCrawler, url: str, run_config: CrawlerRunConfig) -> tuple[str, str]:
    result = await crawler.arun(url=url, config=run_config)
    if not getattr(result, "success", False):
        raise RuntimeError(getattr(result, "error_message", None) or "Crawl4AI reported an unsuccessful crawl.")

    markdown_text, markdown_type = get_raw_markdown(result)
    if not markdown_text.strip():
        raise RuntimeError("Crawl succeeded but returned empty markdown.")
    return markdown_text, markdown_type


def save_markdown_file(markdown_text: str, output_dir: Path, url: str, row_number: int) -> Path:
    domain, filename = slugify_filename_from_url(url, row_number)
    target_dir = output_dir / domain
    target_dir.mkdir(parents=True, exist_ok=True)
    file_path = target_dir / filename
    file_path.write_text(markdown_text, encoding="utf-8")
    return file_path


def update_worksheet_row(worksheet, row_number: int, headers: dict[str, int], **values) -> None:
    for key, value in values.items():
        worksheet.cell(row=row_number, column=headers[normalize_header(key)], value=value)


def relative_to_workbook(excel_path: Path, file_path: Path) -> str:
    return Path(os.path.relpath(file_path, start=excel_path.parent)).as_posix()


async def process_rows(args: argparse.Namespace) -> None:
    excel_path = Path(args.excel_path).resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel workbook not found: {excel_path}")

    output_dir = resolve_output_dir(excel_path, args.output_dir)
    workbook, worksheets = load_workbook_and_sheets(excel_path, args.sheet_name)
    browser_config = BrowserConfig(headless=args.headless)
    run_config = build_run_config()
    summary = {"checked": 0, "success": 0, "failed": 0, "skipped": 0}

    async with AsyncWebCrawler(config=browser_config) as crawler:
        for worksheet in worksheets:
            headers = ensure_output_columns_exist(worksheet, read_header_map(worksheet))
            url_key = detect_url_column(headers, args.url_column)
            url_col = headers[url_key]
            workbook.save(excel_path)
            print(f"\n[SHEET] {worksheet.title} | url_column={url_key}")

            for row_number in range(2, worksheet.max_row + 1):
                url = str(worksheet.cell(row=row_number, column=url_col).value or "").strip()
                if not url:
                    continue

                summary["checked"] += 1
                saved_file = existing_markdown_file(
                    excel_path,
                    worksheet.cell(row=row_number, column=headers["markdown_path"]).value,
                )

                if saved_file and not args.force_rescrape:
                    summary["skipped"] += 1
                    print(f"[SKIP] Row {row_number}: {url}")
                    workbook.save(excel_path)
                    continue

                try:
                    print(f"[CRAWL] Row {row_number}: {url}")
                    markdown_text, markdown_type = await crawl_one_url(crawler, url, run_config)
                    markdown_file = save_markdown_file(markdown_text, output_dir, url, row_number)
                    update_worksheet_row(
                        worksheet,
                        row_number,
                        headers,
                        markdown_path=relative_to_workbook(excel_path, markdown_file),
                        scrape_status="success",
                        scraped_at=datetime.now().isoformat(timespec="seconds"),
                        scrape_error="",
                        markdown_type=markdown_type,
                    )
                    summary["success"] += 1
                except Exception as exc:
                    update_worksheet_row(
                        worksheet,
                        row_number,
                        headers,
                        markdown_path="",
                        scrape_status="failed",
                        scraped_at=datetime.now().isoformat(timespec="seconds"),
                        scrape_error=str(exc),
                        markdown_type="",
                    )
                    summary["failed"] += 1
                    print(f"[FAIL] Row {row_number}: {url} -> {exc}")
                finally:
                    workbook.save(excel_path)

    print("\nSummary")
    print(f"total rows checked: {summary['checked']}")
    print(f"successful scrapes: {summary['success']}")
    print(f"failed scrapes: {summary['failed']}")
    print(f"skipped rows: {summary['skipped']}")


def main() -> None:
    args = parse_args()
    asyncio.run(process_rows(args))


if __name__ == "__main__":
    main()
